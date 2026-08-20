import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_roadmap_workbook():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="1E293B")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="64748B")
    section_font = Font(name=font_family, size=12, bold=True, color="0F172A")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=9, color="1E293B")
    bold_data_font = Font(name=font_family, size=9, bold=True, color="1E293B")
    code_font = Font(name="Consolas", size=9, color="0F172A")
    
    # Fills
    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    blue_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    dark_slate_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Status Fills
    completed_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Emerald
    completed_font = Font(name=font_family, size=9, bold=True, color="166534")
    
    in_progress_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Amber
    in_progress_font = Font(name=font_family, size=9, bold=True, color="92400E")
    
    planned_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid") # Sky
    planned_font = Font(name=font_family, size=9, bold=True, color="075985")
    
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    medium_border_side = Side(border_style="medium", color="94A3B8")
    double_border_side = Side(border_style="double", color="64748B")
    
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=medium_border_side, bottom=medium_border_side)
    total_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=double_border_side)
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def auto_fit_columns(ws, max_widths=None):
        ws.views.sheetView[0].showGridLines = True
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ('$' in cell.number_format or '%' in cell.number_format):
                    val += '    '
                max_len = max(max_len, len(val))
            adjusted_width = max(max_len + 3, 10)
            if max_widths and col_letter in max_widths:
                adjusted_width = min(adjusted_width, max_widths[col_letter])
            ws.column_dimensions[col_letter].width = adjusted_width

    # ==========================================
    # SHEET 1: EXECUTIVE SUMMARY
    # ==========================================
    ws1 = wb.create_sheet(title="1. Executive Summary")
    
    ws1["A1"] = "MONEY CARD PLATFORM — COMPREHENSIVE ROADMAP (3-4 WEEKS TIMELINE)"
    ws1["A1"].font = title_font
    ws1["A2"] = "Source of Truth: M0 V10 Frozen Contract • Standard: 8 Productive Hours / Day • 5 Days / Week • Horizon: 3 to 4 Calendar Weeks"
    ws1["A2"].font = subtitle_font
    
    ws1["A4"] = "1. Overall Project Status & Timeline Summary (3-4 Weeks Aligned)"
    ws1["A4"].font = section_font
    
    headers_s1 = [
        "Platform / Track", "Lead Developer", "Current Architecture Status", 
        "Completed Work (Days)", "Completed Work (Hours)", "Remaining Work (Days)", 
        "Remaining Work (Hours)", "Total Effort (Days)", "Total Effort (Hours)", "Progress (%)"
    ]
    
    for col_idx, h in enumerate(headers_s1, 1):
        cell = ws1.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = header_align
        cell.border = header_border
    ws1.row_dimensions[5].height = 26

    summary_data = [
        ["Backend API (Node.js / Prisma / PostgreSQL)", "Developer 2", "Starting from Scratch (Paced for 3-4 Weeks)", 0, 0, 15, 120, 15, 120, "0.0%"],
        ["Web Admin Frontend (React / TypeScript / Vite)", "Developer 1", "91.3% Complete (M1–M11 Verified)", 21, 168, 1, 8, 22, 176, "95.5%"],
        ["Staff Mobile App (Flutter Native / Riverpod)", "Developer 1", "100% Client Ready (M13–M17, M19 Done)", 11, 88, 1.5, 12, 12.5, 100, "88.0%"],
        ["Joint Post-Merge Integration & Production Release", "Dev 1 + Dev 2", "Scheduled Week 4 (Days 16–20)", 0, 0, 5, 40, 5, 40, "0.0%"],
    ]

    for row_idx, r in enumerate(summary_data, 6):
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(r, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_data_font if col_idx in [1, 2, 10] else data_font
            cell.fill = fill
            cell.border = cell_border
            if col_idx in [4, 5, 6, 7, 8, 9]:
                cell.alignment = right_align
            elif col_idx in [10]:
                cell.alignment = center_align
                cell.fill = completed_fill if "100" in val or "95" in val else in_progress_fill
                cell.font = completed_font if "100" in val or "95" in val else in_progress_font
            else:
                cell.alignment = left_align
        ws1.row_dimensions[row_idx].height = 22

    # Total row
    totals_s1 = ["TOTAL COMBINED PROJECT EFFORT", "All Tracks", "Parallel Execution Schedule", "=SUM(D6:D9)", "=SUM(E6:E9)", "=SUM(F6:F9)", "=SUM(G6:G9)", "=SUM(H6:H9)", "=SUM(I6:I9)", "=AVERAGE(D6/H6, D7/H7, D8/H8, D9/H9)"]
    row_idx = 10
    for col_idx, val in enumerate(totals_s1, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = bold_data_font
        cell.fill = total_fill
        cell.border = total_border
        cell.alignment = right_align if col_idx in [4, 5, 6, 7, 8, 9] else (center_align if col_idx == 10 else left_align)
        if col_idx == 10:
            cell.number_format = '0.0%'
    ws1.row_dimensions[10].height = 24

    # Strategic Highlights Box
    ws1["A12"] = "2. Strategic 3-4 Weeks Schedule Alignment Comparison"
    ws1["A12"].font = section_font
    
    headers_comp = ["Comparison Dimension", "Aggressive 7-Day Fast-Track (Previous)", "Realistic 3-4 Weeks Horizon (Aligned)", "Benefits to Engineering & Delivery Quality"]
    for col_idx, h in enumerate(headers_comp, 1):
        cell = ws1.cell(row=13, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = dark_slate_fill
        cell.alignment = header_align
        cell.border = header_border
    ws1.row_dimensions[13].height = 24

    comp_rows = [
        ["Backend Development Duration", "7 Working Days (56 Hours)", "15 Working Days (120 Hours)", "Ample time to write robust Prisma models, migrations, indexes, and custom business validation."],
        ["Pacing & Daily Workload", "3.0 Milestones / Day (High Risk)", "1.25 to 1.5 Milestones / Day (Sustainable)", "Avoids developer burnout, allows thorough local integration testing of all 65 endpoints."],
        ["Concurrency & Financial Safety", "Basic atomic operations", "Full PostgreSQL $transaction locks & Isolation", "Guarantees zero negative stock, strict balance idempotency, and audit trail safety."],
        ["Joint Integration Window", "3 Working Days (24 Hours)", "5 Working Days (40 Hours)", "Full 1-week dedicated buffer for live CORS, PostgreSQL load testing, and Android QR hardware QA."],
        ["Total Calendar Horizon", "2 Calendar Weeks (10 Days)", "4 Calendar Weeks (20 Working Days)", "Delivers an enterprise-grade, thoroughly tested system ready for immediate live production launch."]
    ]
    
    for row_idx, r in enumerate(comp_rows, 14):
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(r, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_data_font if col_idx == 1 else data_font
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = left_align
        ws1.row_dimensions[row_idx].height = 26

    auto_fit_columns(ws1, {"A": 35, "B": 25, "C": 35, "D": 50})

    # ==========================================
    # SHEET 2: DEVELOPER 2 (BACKEND ROADMAP)
    # ==========================================
    ws2 = wb.create_sheet(title="2. Dev 2 (Backend Roadmap)")
    
    ws2["A1"] = "DEVELOPER 2 (BACKEND) — 3-4 WEEKS IMPLEMENTATION ROADMAP"
    ws2["A1"].font = title_font
    ws2["A2"] = "Stack: Node.js • TypeScript • Express • Prisma ORM • PostgreSQL • Target: 20 Milestones / 65 Routes / 120 Hours (15 Working Days)"
    ws2["A2"].font = subtitle_font
    
    headers_s2 = [
        "Milestone", "Task & Module Name", "Week", "Allocated Days", "Allocated Hours", 
        "Dependencies", "M0 V10 Endpoint Scope", "Key Technical Deliverables & Architecture Scope", "Testing & Verification Criteria"
    ]
    
    for col_idx, h in enumerate(headers_s2, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = header_align
        cell.border = header_border
    ws2.row_dimensions[4].height = 26

    backend_milestones = [
        ["D2-M1", "Backend Foundation, Express & Error Architecture", "Week 1 (Day 1)", 0.5, 4.0, "None", "Global Router", "Express TypeScript bootstrap, .env config, Winston logger, M0 V10 {success, data, error} envelope middleware, CORS.", "Server boots on port 8080/4000, returns formatted M0 error envelope on 404/500."],
        ["D2-M2", "PostgreSQL Database Schema & Prisma ORM Setup", "Week 1 (Day 1-2)", 0.75, 6.0, "D2-M1", "DB Migrations", "12 Prisma models (Organization, Branch, User, Card, CardSession, Product, InventoryItem, InventoryMovement, Transaction, Plan, PlanRequest, AuditLog). Seed from TEST_DATA.json.", "Prisma migrate deploy succeeds, seed script populates demo organizations, branches, and staff."],
        ["D2-M3", "Authentication Module & JWT Lifecycle", "Week 1 (Day 2)", 0.75, 6.0, "D2-M2", "POST /auth/login, POST /auth/refresh, POST /auth/logout, GET /auth/me, POST /auth/forgot-password, POST /auth/reset-password, POST /auth/change-password (7 routes)", "Argon2/Bcrypt password hashing, 15m JWT access tokens, 7d refresh tokens with database revocation, getMe enriched with permissions & branches.", "Login returns valid tokens, expired access token refreshes seamlessly, logout revokes refresh token."],
        ["D2-M4", "Authorization, RBAC & Multi-Tenant Middleware", "Week 1 (Day 3)", 0.75, 6.0, "D2-M3", "Middleware Layer", "Role guards (SUPER_ADMIN, ORG_ADMIN, STAFF), 20 granular AppPermission evaluators, automatic tenant organization & assigned branch isolation filter.", "Unauthorized requests return 401, missing permission returns 403 FORBIDDEN, cross-org access blocked."],
        ["D2-M5", "Organizations Management APIs (Super Admin)", "Week 1 (Day 4)", 0.75, 6.0, "D2-M4", "GET /organizations, POST /organizations, GET /organizations/:id, PUT /organizations/:id, DELETE /organizations/:id (5 routes)", "Super Admin tenant listing, tenant provisioning with default branch & admin creation, status updates with cascading deactivation.", "Super Admin can create tenant, edit status; regular staff cannot access /organizations."],
        ["D2-M6", "Subscription Plans, Limits & Quota Overrides", "Week 1 (Day 5)", 0.75, 6.0, "D2-M5", "GET /plans, POST /plans, PUT /plans/:id, GET /plans/requests, POST /plans/requests, POST /plans/requests/:id/approve, POST /plans/requests/:id/reject, PUT /organizations/:id/limits (8 routes)", "Global pricing plans (NO transaction limits), custom org quota overrides (card/branch limits), plan change approval workflow.", "Plan limits enforced when org exceeds max branches/cards; approval upgrades tenant plan."],
        ["D2-M7", "Branches & Staff Management APIs", "Week 2 (Day 6)", 0.75, 6.0, "D2-M4", "GET /branches, POST /branches, GET /branches/:id, PUT /branches/:id, GET /staff, POST /staff, GET /staff/:id, PUT /staff/:id, DELETE /staff/:id (9 routes)", "Branch CRUD with org isolation, Staff CRUD, branch assignments replacement, permission array replacement, branch count limit checks.", "Staff assigned to Branch 1 cannot access Branch 2; branch creation respects plan limits."],
        ["D2-M8", "Card Inventory & Dual QR Resolution Engine", "Week 2 (Day 7)", 0.75, 6.0, "D2-M7", "GET /cards, POST /cards, GET /cards/:id, POST /cards/resolve, POST /cards/:id/block, POST /cards/:id/unblock (6 routes)", "Authoritative card inventory, AVAILABLE state on import, opaque qrToken mapping, POST /cards/resolve with active session enrichment, block/unblock.", "QR resolve returns 404 for unknown QR, returns card + active session for valid QR, block prevents operations."],
        ["D2-M9", "Card Sessions Lifecycle & Balance Management", "Week 2 (Day 8)", 0.75, 6.0, "D2-M8", "GET /card-sessions, POST /card-sessions, GET /card-sessions/:id (3 routes)", "Session creation on AVAILABLE card (transitions card to ACTIVE), active session list filtered by branch, session detail with balance.", "Cannot create session on ACTIVE/BLOCKED card; active session count dynamically accurate."],
        ["D2-M10", "Product Catalog Management APIs", "Week 2 (Day 9)", 0.75, 6.0, "D2-M7", "GET /products, POST /products, GET /products/:id, PUT /products/:id, DELETE /products/:id (5 routes)", "Branch product catalog, multi-select category array (List<String>, NO tags), price validation, auto InventoryItem record initialization.", "Products correctly scoped to branch; category array saves and filters accurately."],
        ["D2-M11", "POS Checkout & Purchase Financials Engine", "Week 2 (Day 10)", 1.0, 8.0, "D2-M9, D2-M10", "POST /card-sessions/:id/purchase (1 route)", "Atomic Prisma $transaction: server price lookup, balance verification, stock verification, balance deduction, stock decrease, Transaction record.", "Insufficient balance returns 400; concurrency test ensures zero double-spending or negative balances."],
        ["D2-M12", "Card Recharge & Payment Records", "Week 3 (Day 11)", 0.75, 6.0, "D2-M9", "POST /card-sessions/:id/recharge (1 route)", "Card session top-up with CASH or manual UPI verification reference, atomic balance credit, idempotency reference validation, Transaction record.", "Recharge credits session balance immediately; duplicate externalReference is rejected."],
        ["D2-M13", "Card Return, Refund & Session Settlement", "Week 3 (Day 11-12)", 0.75, 6.0, "D2-M11, D2-M12", "POST /card-sessions/:id/return (1 route)", "Session settlement: calculates remaining balance, records REFUND transaction, zeroes session balance, marks session SETTLED, resets card to AVAILABLE.", "Settled session cannot be reused; card immediately available for re-issuance with ₹0 balance."],
        ["D2-M14", "Inventory Stock Control & 3-Column CSV Import", "Week 3 (Day 12-13)", 1.0, 8.0, "D2-M10", "GET /inventory, POST /inventory/:id/adjust, GET /inventory/movements, POST /inventory/import-csv, GET /inventory/csv-template (5 routes)", "Stock level tracking, manual stock adjustments with audit reason (RESTOCK/ADJUSTMENT), 3-col CSV parser (itemName,category,price) with 2-phase preview & commit.", "Stock adjustment logs InventoryMovement; CSV import previews rows and creates products + inventory."],
        ["D2-M15", "Real-Time Branch Analytics & Aggregations", "Week 3 (Day 13)", 0.75, 6.0, "D2-M11, D2-M12", "GET /analytics (1 route)", "Aggregated metrics calculation: total revenue, recharge volume, purchase volume, active sessions, peak demand hours, top-selling items.", "Analytics metrics match sum of completed transactions for requested branch and date range."],
        ["D2-M16", "Formal PDF Reports Engine (Binary Streaming)", "Week 3 (Day 14)", 0.75, 6.0, "D2-M15", "GET /reports, POST /reports/generate, GET /reports/:id/download (3 routes)", "Formal business reports list, asynchronous report generator using PDFKit, binary streaming PDF download conforming to M0 V10.", "GET /reports/:id/download returns application/pdf with valid binary PDF content."],
        ["D2-M17", "Public Customer User Portal APIs", "Week 3 (Day 14)", 0.75, 6.0, "D2-M9", "POST /portal/resolve, GET /portal/session, GET /portal/transactions (3 routes)", "Public card QR resolution with data masking, temporary portal session token, live card balance lookup, customer transaction receipt history.", "Public user can check balance and view digital receipts without staff credentials."],
        ["D2-M18", "System Audit Logging & Interceptors", "Week 3 (Day 15)", 0.75, 6.0, "D2-M4", "GET /audit-logs (1 route)", "Automated Prisma middleware logging all auth events, card state transitions, financial transactions, staff changes, and plan modifications.", "Sensitive operations automatically log actor ID, action, entity, and timestamp to AuditLog table."],
        ["D2-M19", "API Contract Parity Verification (All 65 Routes)", "Week 3 (Day 15)", 0.75, 6.0, "D2-M1 to D2-M18", "Full 65 Routes", "Rigorous schema validation verifying every endpoint against MONEY_CARD_COMPLETE_API_CONTRACT_V10.pdf: envelopes, status codes, query params.", "100% route parity verified; zero schema deviations from M0 V10 specification."],
        ["D2-M20", "Postman QA Collection Pass & Handover", "Week 3 (Day 15)", 0.75, 6.0, "D2-M19", "Postman Suite (65 Requests)", "Automated Postman Collection Runner execution covering all 65 endpoints across Super Admin, Org Admin, and Staff roles.", "65/65 Postman assertions pass green (100% success rate); backend certified ready for frontend merge."]
    ]

    for row_idx, r in enumerate(backend_milestones, 5):
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(r, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_data_font if col_idx in [1, 2] else (code_font if col_idx == 7 else data_font)
            cell.fill = fill
            cell.border = cell_border
            if col_idx in [4, 5]:
                cell.alignment = right_align
            elif col_idx in [1, 3, 6]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        ws2.row_dimensions[row_idx].height = 28

    # Total Row
    tot_r = len(backend_milestones) + 5
    ws2.cell(row=tot_r, column=1, value="DEVELOPER 2 TOTAL (BACKEND TRACK)").font = bold_data_font
    ws2.cell(row=tot_r, column=2, value="All 20 Milestones / 65 Authoritative Routes").font = bold_data_font
    ws2.cell(row=tot_r, column=3, value="Weeks 1 to 3").font = bold_data_font
    ws2.cell(row=tot_r, column=4, value="=SUM(D5:D24)").font = bold_data_font
    ws2.cell(row=tot_r, column=5, value="=SUM(E5:E24)").font = bold_data_font
    ws2.cell(row=tot_r, column=6, value="Full Pipeline").font = bold_data_font
    ws2.cell(row=tot_r, column=7, value="65/65 Endpoints").font = bold_data_font
    ws2.cell(row=tot_r, column=8, value="Complete PostgreSQL & Node.js Production Backend").font = bold_data_font
    ws2.cell(row=tot_r, column=9, value="Postman 100% Pass Certification").font = bold_data_font

    for col_idx in range(1, 10):
        cell = ws2.cell(row=tot_r, column=col_idx)
        cell.fill = total_fill
        cell.border = total_border
        if col_idx in [4, 5]:
            cell.alignment = right_align

    auto_fit_columns(ws2, {"A": 12, "B": 32, "C": 18, "D": 14, "E": 14, "F": 16, "G": 40, "H": 50, "I": 45})

    # ==========================================
    # SHEET 3: DEVELOPER 1 (WEB FRONTEND)
    # ==========================================
    ws3 = wb.create_sheet(title="3. Dev 1 (Web Frontend)")
    
    ws3["A1"] = "DEVELOPER 1 (WEB FRONTEND) — ROADMAP & STATUS"
    ws3["A1"].font = title_font
    ws3["A2"] = "Stack: React 18 • TypeScript • Vite • Tailwind / Custom CSS • Status: 91.3% Complete (M1–M11 Verified, M12 Pending Backend)"
    ws3["A2"].font = subtitle_font
    
    headers_s3 = ["Milestone", "Task Name", "Time (Days)", "Time (Hours)", "Current Status", "Key Scope & Technical Deliverables"]
    for col_idx, h in enumerate(headers_s3, 1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = dark_slate_fill
        cell.alignment = header_align
        cell.border = header_border
    ws3.row_dimensions[4].height = 26

    web_milestones = [
        ["M1", "Foundation & Design System", 2.0, 16.0, "COMPLETED", "Design tokens, dark/light themes, responsive navigation, top header, breadcrumbs, standard UI kit."],
        ["M2", "Authentication & RBAC Suite", 2.0, 16.0, "COMPLETED", "Login screens, forgot password modal (neutral response), reset password, change password, profile dropdown, role guards."],
        ["M3", "Cards Management & Issue", 2.0, 16.0, "COMPLETED", "Card list, QR barcode viewer modal, manual card creation, CSV template download & import, status filter with Available count."],
        ["M4", "Sessions Management", 1.75, 14.0, "COMPLETED", "Active session dashboard, session issuance, branch filtering, live session balance indicators, session details modal."],
        ["M5", "POS Terminal & Products", 2.25, 18.0, "COMPLETED", "POS checkout cart, product catalog grid, search/filter, multi-select category tags (string[]), price calculation (no tags)."],
        ["M6", "Payments, Recharge & Refunds", 1.5, 12.0, "COMPLETED", "Cash top-up modal, manual UPI verification modal, digital receipt modal, session return/refund modal with settlement confirmation."],
        ["M7", "Inventory & 3-Column CSV", 2.0, 16.0, "COMPLETED", "Live inventory table, stock adjustments, low stock badges, 3-column CSV template download, branch 2-phase preview & commit."],
        ["M8", "Analytics & Peak Demand", 2.0, 16.0, "COMPLETED", "Real-time analytics dashboard, branch performance comparison table, peak hours traffic charts, hourly breakdown, PDF export."],
        ["M9", "Offline Resilience & Mock Store", 1.5, 12.0, "COMPLETED", "In-memory mock database store, realistic latency simulation, localStorage state persistence, reset store utility."],
        ["M10", "Admin Controls & Subscriptions", 2.25, 18.0, "COMPLETED", "Super Admin tenant org manager, plan comparison matrix, plan upgrade/downgrade request form, limit overrides manager."],
        ["M11", "Web Polish & Contract QA", 1.75, 14.0, "COMPLETED", "Automated contract test runner integration, 58/58 test pass verification, zero TypeScript errors, production build optimization."],
        ["M12", "Web Real Backend Integration", 1.0, 8.0, "REMAINING (Week 4)", "Switch .env VITE_USE_MOCK_API=false to point to live Node.js server (http://localhost:4000/api/v1), smoke test all 65 routes."]
    ]

    for row_idx, r in enumerate(web_milestones, 5):
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(r, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_data_font if col_idx in [1, 2] else data_font
            cell.fill = fill
            cell.border = cell_border
            if col_idx in [3, 4]:
                cell.alignment = right_align
            elif col_idx == 5:
                cell.alignment = center_align
                cell.fill = completed_fill if "COMPLETED" in val else in_progress_fill
                cell.font = completed_font if "COMPLETED" in val else in_progress_font
            else:
                cell.alignment = left_align
        ws3.row_dimensions[row_idx].height = 24

    auto_fit_columns(ws3, {"A": 12, "B": 30, "C": 14, "D": 14, "E": 22, "F": 55})

    # ==========================================
    # SHEET 4: DEVELOPER 1 (FLUTTER MOBILE APP)
    # ==========================================
    ws4 = wb.create_sheet(title="4. Dev 1 (Flutter Staff App)")
    
    ws4["A1"] = "DEVELOPER 1 (FLUTTER STAFF MOBILE APP) — ROADMAP & STATUS"
    ws4["A1"].font = title_font
    ws4["A2"] = "Stack: Flutter 3.x • Dart • Riverpod 2.6 • Dio • Mobile Scanner • EscPos Thermal • Status: M13–M17, M19 Completed (118 Tests Passing)"
    ws4["A2"].font = subtitle_font
    
    headers_s4 = ["Milestone", "Task Name", "Time (Days)", "Time (Hours)", "Current Status", "Key Scope & Technical Deliverables"]
    for col_idx, h in enumerate(headers_s4, 1):
        cell = ws4.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = dark_slate_fill
        cell.alignment = header_align
        cell.border = header_border
    ws4.row_dimensions[4].height = 26

    flutter_milestones = [
        ["M13", "Flutter Foundation & Architecture", 1.75, 14.0, "COMPLETED", "Light theme, Dio network client with Auth/Error/Mock interceptors, SecureStorage token manager, AppRouter, AppShell."],
        ["M14", "Staff Authentication & Permissions", 1.5, 12.0, "COMPLETED", "Login screen, token refresh handling, branch switcher, PermissionGuard, role/permission providers, auth service."],
        ["M15", "Staff Cards & QR Operations", 2.25, 18.0, "COMPLETED", "Camera QR scanner (QRScannerView), manual card entry fallback, QR resolve (POST /cards/resolve), card list, Home active session counter."],
        ["M16", "Payments, POS Checkout & Return", 2.75, 22.0, "COMPLETED", "Cash recharge screen, manual UPI verification screen (store physical QR used), POS catalog cart (POSCartProvider), return card & refund screen."],
        ["M17", "Inventory & Shift Analytics", 2.0, 16.0, "COMPLETED", "Inventory screen with live stock, quick stock adjustment, AddProductScreen with multi-category selector, staff AnalyticsScreen with volume cards."],
        ["M18", "Real Backend Integration & QA Audit", 1.0, 8.0, "AUDIT READY (Week 4)", "Integration readiness audit passed (FLUTTER_REAL_API_INTEGRATION_AUDIT.md). Switch AppConfig.apiMode = real to point to live backend."],
        ["M19", "Hardware, Scanner & Mock QR Hardening", 1.0, 8.0, "COMPLETED", "Camera torch toggle, scan vibration feedback, ESC/POS thermal receipt formatting (58mm/80mm), physical camera Mock QR display."],
        ["M20", "Android Production QA & Release", 0.5, 4.0, "REMAINING (Week 4)", "Release keystore signing, ProGuard rules, Android APK/AAB build generation, offline network error handling validation."]
    ]

    for row_idx, r in enumerate(flutter_milestones, 5):
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(r, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_data_font if col_idx in [1, 2] else data_font
            cell.fill = fill
            cell.border = cell_border
            if col_idx in [3, 4]:
                cell.alignment = right_align
            elif col_idx == 5:
                cell.alignment = center_align
                cell.fill = completed_fill if "COMPLETED" in val else in_progress_fill
                cell.font = completed_font if "COMPLETED" in val else in_progress_font
            else:
                cell.alignment = left_align
        ws4.row_dimensions[row_idx].height = 24

    auto_fit_columns(ws4, {"A": 12, "B": 32, "C": 14, "D": 14, "E": 24, "F": 55})

    # ==========================================
    # SHEET 5: MASTER PARALLEL SCHEDULE (3-4 WEEKS)
    # ==========================================
    ws5 = wb.create_sheet(title="5. Master Daily Schedule (3-4W)")
    
    ws5["A1"] = "MASTER PARALLEL DAILY SCHEDULE — 4 WEEKS (20 WORKING DAYS)"
    ws5["A1"].font = title_font
    ws5["A2"] = "Coordinated Multi-Track Sprint Plan: Backend Build (Weeks 1–3) + Joint Integration & Release QA (Week 4)"
    ws5["A2"].font = subtitle_font
    
    headers_s5 = [
        "Calendar Timeline", "Developer 2 (Backend Track)", "Developer 1 (Web & Flutter Tracks)", 
        "Daily Target Deliverable & Milestone Output", "Sign-Off Status"
    ]
    
    for col_idx, h in enumerate(headers_s5, 1):
        cell = ws5.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = header_align
        cell.border = header_border
    ws5.row_dimensions[4].height = 26

    daily_schedule = [
        # Week 1
        ["Week 1 • Day 1", "D2-M1: Express TS Setup & Error Envelope Middleware", "Flutter Mock QR Codes Screen & Scanner Validation", "Backend server running with global error envelope; Flutter camera test verified.", "PLANNED"],
        ["Week 1 • Day 2", "D2-M2: PostgreSQL Connection & Prisma Schema Models", "Web Frontend M1–M6 Contract Review & Mock Sync", "12 Prisma database models created & migrated; seed data loaded.", "PLANNED"],
        ["Week 1 • Day 3", "D2-M3: Authentication Module (JWT, Refresh, Password)", "Flutter Hardware Settings & Thermal Printer Hardening", "Login, refresh, and logout APIs fully operational with Argon2 hashing.", "PLANNED"],
        ["Week 1 • Day 4", "D2-M4: RBAC & Multi-Tenant Isolation Middleware", "Web Frontend M7–M10 Tenancy & RBAC Verification", "Role guards and organization/branch tenant scoping active on all routes.", "PLANNED"],
        ["Week 1 • Day 5", "D2-M5: Organizations & D2-M6: Plans Management APIs", "Web & Flutter Shared Model Type-Check against M0 V10", "Super Admin tenant listing, provisioning, and global plan limit APIs complete.", "PLANNED"],
        
        # Week 2
        ["Week 2 • Day 6", "D2-M7: Branches & Staff Management CRUD APIs", "Flutter Active Sessions Dynamic Calculation QA", "Branch CRUD, staff permissions replacement, and branch limit validation live.", "PLANNED"],
        ["Week 2 • Day 7", "D2-M8: Card Inventory & Dual QR Resolution APIs", "Flutter Card Resolver & QR Debouncing Verification", "Authoritative card inventory with AVAILABLE status & QR token resolution live.", "PLANNED"],
        ["Week 2 • Day 8", "D2-M9: Card Sessions Lifecycle & Active Balances", "Flutter Issue Card Session Creation Flow Testing", "Session creation on available cards & active session list endpoints live.", "PLANNED"],
        ["Week 2 • Day 9", "D2-M10: Product Catalog APIs (Multi-Category Array)", "Web POS Catalog Multi-Select Category Verification", "Branch product CRUD with category arrays (zero tags) & auto-inventory init.", "PLANNED"],
        ["Week 2 • Day 10", "D2-M11: POS Purchase Engine ($transaction Locks)", "Flutter POS Checkout Cart & Decimal Price Verification", "Atomic purchase execution with server price lookup & stock deduction live.", "PLANNED"],
        
        # Week 3
        ["Week 3 • Day 11", "D2-M12: Recharge & D2-M13: Session Refund/Return", "Flutter Recharge & Return Settlement Dialogs QA", "CASH/UPI wallet recharge and session refund/settlement transactions live.", "PLANNED"],
        ["Week 3 • Day 12", "D2-M14: Inventory Stock Adjustments & CSV Import", "Web 3-Column CSV Import & Live Inventory Testing", "Inventory adjustments with audit logs & 3-column CSV import engine live.", "PLANNED"],
        ["Week 3 • Day 13", "D2-M15: Real-Time Analytics & Aggregated Metrics", "Web & Flutter Analytics Volume Cards Parity Check", "Real-time revenue, recharge, purchase, and peak hour analytics live.", "PLANNED"],
        ["Week 3 • Day 14", "D2-M16: Reports Engine (PDF) & D2-M17: Customer Portal", "Flutter Digital Receipt Preview Monospace QA", "Binary PDF streaming reports download & public customer portal APIs live.", "PLANNED"],
        ["Week 3 • Day 15", "D2-M18: Audit Logging & D2-M19/M20: Postman QA Pass", "Full 65-Route Contract Verification Sign-Off", "All 65 routes passing Postman Collection 100%; BACKEND DELIVERED FOR MERGE!", "MILESTONE GATE"],
        
        # Week 4 (Joint Integration & Production Release)
        ["Week 4 • Day 16", "Joint Backend Support: CORS, Network Latency, DB Logs", "Web M12: Real API Integration (65 Routes E2E QA)", "Web Admin Frontend 100% connected to live backend & PostgreSQL.", "JOINT QA"],
        ["Week 4 • Day 17", "Joint Backend Support: JWT Refresh & Session Monitoring", "Flutter M18: Real API Integration & Live POS QA", "Flutter Staff App 100% connected to live backend; live card issuing verified.", "JOINT QA"],
        ["Week 4 • Day 18", "Joint Concurrency & Multi-Tenant Boundary Stress Testing", "Simultaneous Multi-Branch Purchase & Recharge Testing", "Database locks verified under concurrent purchasing; zero data leakage.", "JOINT QA"],
        ["Week 4 • Day 19", "Final Security Audit, Input Sanitization & SQL Guard", "Flutter M19 Hardware & ESC/POS Thermal Printing Live Test", "Hardware printers and physical scanner tested against live database.", "JOINT QA"],
        ["Week 4 • Day 20", "Production Database Migration & Server Deployment", "Flutter M20: Keystore Signing & Release APK/AAB Build", "FINAL PRODUCTION SIGN-OFF: Web, Mobile, and Backend 100% Live!", "FINAL RELEASE"]
    ]

    for row_idx, r in enumerate(daily_schedule, 5):
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        is_week_end = row_idx in [9, 14, 19, 24]
        for col_idx, val in enumerate(r, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_data_font if col_idx in [1, 5] else data_font
            cell.fill = fill
            cell.border = cell_border
            if col_idx == 1:
                cell.alignment = center_align
            elif col_idx == 5:
                cell.alignment = center_align
                if val == "FINAL RELEASE":
                    cell.fill = completed_fill
                    cell.font = completed_font
                elif val == "MILESTONE GATE":
                    cell.fill = in_progress_fill
                    cell.font = in_progress_font
                else:
                    cell.fill = planned_fill
                    cell.font = planned_font
            else:
                cell.alignment = left_align
        ws5.row_dimensions[row_idx].height = 24

    auto_fit_columns(ws5, {"A": 18, "B": 45, "C": 45, "D": 50, "E": 20})

    # ==========================================
    # SHEET 6: 65-ROUTE API SPECIFICATION MATRIX
    # ==========================================
    ws6 = wb.create_sheet(title="6. 65-Route API Matrix")
    
    ws6["A1"] = "M0 V10 COMPLETE 65-ROUTE AUTHORITATIVE API SPECIFICATION MATRIX"
    ws6["A1"].font = title_font
    ws6["A2"] = "Frozen Shared System Contract • Standard: /api/v1 Base Path • Envelope: {success: boolean, data: object, error: object}"
    ws6["A2"].font = subtitle_font
    
    headers_s6 = ["Route #", "Module Domain", "HTTP Method", "API Endpoint Path", "Required Role / Permission", "Request Payload / Parameters", "Authoritative M0 V10 Response"]
    for col_idx, h in enumerate(headers_s6, 1):
        cell = ws6.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = header_align
        cell.border = header_border
    ws6.row_dimensions[4].height = 26

    api_routes = [
        # Auth (1-7)
        [1, "Authentication", "POST", "/api/v1/auth/login", "Public", "{email, password}", "{accessToken, refreshToken, user: {id, email, role, permissions}}"],
        [2, "Authentication", "POST", "/api/v1/auth/refresh", "Public (Bearer Refresh)", "{refreshToken}", "{accessToken, refreshToken}"],
        [3, "Authentication", "POST", "/api/v1/auth/logout", "Authenticated", "None", "{message: 'Logged out successfully'}"],
        [4, "Authentication", "GET", "/api/v1/auth/me", "Authenticated", "None", "{user: {id, email, name, role, permissions, assignedBranchIds}}"],
        [5, "Authentication", "POST", "/api/v1/auth/forgot-password", "Public", "{email}", "{message: 'Password reset link sent'}"],
        [6, "Authentication", "POST", "/api/v1/auth/reset-password", "Public", "{token, newPassword}", "{message: 'Password reset successful'}"],
        [7, "Authentication", "POST", "/api/v1/auth/change-password", "Authenticated", "{currentPassword, newPassword}", "{message: 'Password changed successfully'}"],

        # Organizations (8-12)
        [8, "Organizations", "GET", "/api/v1/organizations", "SUPER_ADMIN", "?status, ?search, ?page, ?limit", "{items: [{id, name, status, planId, branchCount}]}"],
        [9, "Organizations", "POST", "/api/v1/organizations", "SUPER_ADMIN", "{name, email, planId, adminName, adminPassword}", "{organization: {id, name, status}, admin: {id, email}}"],
        [10, "Organizations", "GET", "/api/v1/organizations/:id", "SUPER_ADMIN / ORG_ADMIN", "Path: id", "{id, name, status, planId, limits, createdAt}"],
        [11, "Organizations", "PUT", "/api/v1/organizations/:id", "SUPER_ADMIN", "Path: id, {name, status, planId}", "{id, name, status, updatedAt}"],
        [12, "Organizations", "DELETE", "/api/v1/organizations/:id", "SUPER_ADMIN", "Path: id", "{message: 'Organization deactivated'}"],

        # Plans & Subscriptions (13-20)
        [13, "Plans & Subscriptions", "GET", "/api/v1/plans", "Authenticated", "None", "[{id, name, price, maxBranches, maxCards, features}]"],
        [14, "Plans & Subscriptions", "POST", "/api/v1/plans", "SUPER_ADMIN", "{name, price, maxBranches, maxCards, features}", "{id, name, price, maxBranches, maxCards}"],
        [15, "Plans & Subscriptions", "PUT", "/api/v1/plans/:id", "SUPER_ADMIN", "Path: id, {name, price, maxBranches, maxCards}", "{id, name, price, updatedAt}"],
        [16, "Plans & Subscriptions", "GET", "/api/v1/plans/requests", "SUPER_ADMIN", "?status, ?page, ?limit", "{items: [{id, organizationId, requestedPlanId, status}]}"],
        [17, "Plans & Subscriptions", "POST", "/api/v1/plans/requests", "ORG_ADMIN", "{requestedPlanId, reason}", "{id, organizationId, status: 'PENDING'}"],
        [18, "Plans & Subscriptions", "POST", "/api/v1/plans/requests/:id/approve", "SUPER_ADMIN", "Path: id", "{id, status: 'APPROVED', effectivePlanId}"],
        [19, "Plans & Subscriptions", "POST", "/api/v1/plans/requests/:id/reject", "SUPER_ADMIN", "Path: id, {reason}", "{id, status: 'REJECTED', reason}"],
        [20, "Plans & Subscriptions", "PUT", "/api/v1/organizations/:id/limits", "SUPER_ADMIN", "Path: id, {maxBranches, maxCards}", "{organizationId, customLimits: {maxBranches, maxCards}}"],

        # Branches (21-24)
        [21, "Branches", "GET", "/api/v1/branches", "Authenticated", "?organizationId", "[{id, organizationId, name, address, phone, isActive}]"],
        [22, "Branches", "POST", "/api/v1/branches", "ORG_ADMIN (BRANCH_MANAGE)", "{name, address, phone}", "{id, organizationId, name, isActive: true}"],
        [23, "Branches", "GET", "/api/v1/branches/:id", "Authenticated", "Path: id", "{id, organizationId, name, address, phone, isActive}"],
        [24, "Branches", "PUT", "/api/v1/branches/:id", "ORG_ADMIN (BRANCH_MANAGE)", "Path: id, {name, address, phone, isActive}", "{id, name, address, isActive, updatedAt}"],

        # Staff (25-29)
        [25, "Staff Management", "GET", "/api/v1/staff", "ORG_ADMIN (STAFF_VIEW)", "?branchId, ?page, ?limit", "{items: [{id, name, email, role, permissions, assignedBranchIds}]}"],
        [26, "Staff Management", "POST", "/api/v1/staff", "ORG_ADMIN (STAFF_MANAGE)", "{name, email, password, assignedBranchIds, permissions}", "{id, name, email, role: 'STAFF'}"],
        [27, "Staff Management", "GET", "/api/v1/staff/:id", "ORG_ADMIN (STAFF_VIEW)", "Path: id", "{id, name, email, permissions, assignedBranchIds}"],
        [28, "Staff Management", "PUT", "/api/v1/staff/:id", "ORG_ADMIN (STAFF_MANAGE)", "Path: id, {name, assignedBranchIds, permissions, isActive}", "{id, name, permissions, updatedAt}"],
        [29, "Staff Management", "DELETE", "/api/v1/staff/:id", "ORG_ADMIN (STAFF_MANAGE)", "Path: id", "{message: 'Staff user deactivated'}"],

        # Cards (30-35)
        [30, "Cards & Inventory", "GET", "/api/v1/cards", "STAFF (CARD_VIEW)", "?branchId, ?status, ?search, ?page, ?limit", "{items: [{id, qrToken, physicalCardNumber, status, currentBranchId}]}"],
        [31, "Cards & Inventory", "POST", "/api/v1/cards", "STAFF (CARD_ISSUE)", "{physicalCardNumber, branchId}", "{id, qrToken, physicalCardNumber, status: 'AVAILABLE'}"],
        [32, "Cards & Inventory", "GET", "/api/v1/cards/:id", "STAFF (CARD_VIEW)", "Path: id", "{id, qrToken, physicalCardNumber, status, currentBranchId}"],
        [33, "Cards & Inventory", "POST", "/api/v1/cards/resolve", "STAFF (CARD_VIEW)", "{qrToken}", "{card: {id, physicalCardNumber, status}, session: {id, balance}}"],
        [34, "Cards & Inventory", "POST", "/api/v1/cards/:id/block", "STAFF (CARD_BLOCK)", "Path: id, {reason}", "{id, status: 'BLOCKED', blockReason}"],
        [35, "Cards & Inventory", "POST", "/api/v1/cards/:id/unblock", "STAFF (CARD_UNBLOCK)", "Path: id", "{id, status: 'AVAILABLE'}"],

        # Card Sessions (36-41)
        [36, "Card Sessions", "GET", "/api/v1/card-sessions", "STAFF (SESSION_VIEW)", "?branchId, ?status, ?page, ?limit", "{items: [{id, cardId, physicalCardNumber, branchId, status, balance}]}"],
        [37, "Card Sessions", "POST", "/api/v1/card-sessions", "STAFF (SESSION_CREATE)", "{cardId, branchId}", "{id, cardId, branchId, status: 'ACTIVE', balance: 0.0}"],
        [38, "Card Sessions", "GET", "/api/v1/card-sessions/:id", "STAFF (SESSION_VIEW)", "Path: id", "{id, cardId, physicalCardNumber, branchId, status, balance, startedAt}"],
        [39, "Card Sessions", "POST", "/api/v1/card-sessions/:id/purchase", "STAFF (PURCHASE)", "Path: id, {items: [{productId, quantity}]}", "{transactionId, amount, balance, status: 'SUCCESS'}"],
        [40, "Card Sessions", "POST", "/api/v1/card-sessions/:id/recharge", "STAFF (RECHARGE)", "Path: id, {amount, paymentMethod, externalReference}", "{transactionId, amount, balance, paymentMethod, status: 'SUCCESS'}"],
        [41, "Card Sessions", "POST", "/api/v1/card-sessions/:id/return", "STAFF (REFUND)", "Path: id", "{sessionId, refundedAmount, sessionStatus: 'SETTLED', cardStatus: 'AVAILABLE'}"],

        # Products (42-46)
        [42, "Products Catalog", "GET", "/api/v1/products", "STAFF (PRODUCT_VIEW)", "?branchId, ?category, ?status, ?page, ?limit", "{items: [{id, branchId, itemName, category: [], price, status}]}"],
        [43, "Products Catalog", "POST", "/api/v1/products", "ORG_ADMIN (PRODUCT_MANAGE)", "{branchId, itemName, category: [], price, status}", "{id, branchId, itemName, category, price, status: 'ACTIVE'}"],
        [44, "Products Catalog", "GET", "/api/v1/products/:id", "STAFF (PRODUCT_VIEW)", "Path: id", "{id, branchId, itemName, category: [], price, status}"],
        [45, "Products Catalog", "PUT", "/api/v1/products/:id", "ORG_ADMIN (PRODUCT_MANAGE)", "Path: id, {itemName, category: [], price, status}", "{id, itemName, category, price, status, updatedAt}"],
        [46, "Products Catalog", "DELETE", "/api/v1/products/:id", "ORG_ADMIN (PRODUCT_MANAGE)", "Path: id", "{message: 'Product deleted'}"],

        # Inventory (47-51)
        [47, "Inventory Control", "GET", "/api/v1/inventory", "STAFF (INVENTORY_VIEW)", "?branchId, ?search, ?status, ?page, ?limit", "{items: [{id, productId, productName, branchId, currentStock, status}]}"],
        [48, "Inventory Control", "POST", "/api/v1/inventory/:id/adjust", "STAFF (INVENTORY_MANAGE)", "Path: id, {adjustment, reason}", "{id, productId, currentStock, status, updatedAt}"],
        [49, "Inventory Control", "GET", "/api/v1/inventory/movements", "STAFF (INVENTORY_VIEW)", "?branchId, ?inventoryId, ?limit", "{items: [{id, inventoryId, productName, type, quantity, reason}]}"],
        [50, "Inventory Control", "POST", "/api/v1/inventory/import-csv", "ORG_ADMIN (INVENTORY_MANAGE)", "Multipart form-data: file (.csv), ?commit=true/false", "{totalRows, validRows, invalidRows, preview: [], committed: boolean}"],
        [51, "Inventory Control", "GET", "/api/v1/inventory/csv-template", "Authenticated", "None", "CSV file attachment (itemName,category,price)"],

        # Analytics (52-53)
        [52, "Analytics & Metrics", "GET", "/api/v1/analytics", "STAFF (VIEW_ANALYTICS)", "?branchId, ?range (today, week, month)", "{branchPerformance: [{branchId, totalRevenue, rechargeRevenue, purchases}]}"],
        [53, "Analytics & Metrics", "GET", "/api/v1/analytics/demand", "STAFF (VIEW_ANALYTICS)", "?branchId, ?range", "{topProducts: [{productId, itemName, quantitySold, revenue}]}"],

        # Reports (54-56)
        [54, "Reports Engine", "GET", "/api/v1/reports", "ORG_ADMIN (VIEW_REPORTS)", "?branchId, ?type, ?page, ?limit", "{items: [{id, branchId, title, type, dateRange, status}]}"],
        [55, "Reports Engine", "POST", "/api/v1/reports/generate", "ORG_ADMIN (VIEW_REPORTS)", "{branchId, type, startDate, endDate}", "{reportId, status: 'GENERATING'}"],
        [56, "Reports Engine", "GET", "/api/v1/reports/:id/download", "ORG_ADMIN (VIEW_REPORTS)", "Path: id", "Binary PDF file attachment (application/pdf)"],

        # Customer Portal (57-59)
        [57, "Customer Portal", "POST", "/api/v1/portal/resolve", "Public", "{qrToken}", "{portalToken, maskedCardNumber, status}"],
        [58, "Customer Portal", "GET", "/api/v1/portal/session", "Public (Portal Token)", "Header: X-Portal-Token", "{cardId, activeBalance, status, sessionStartedAt}"],
        [59, "Customer Portal", "GET", "/api/v1/portal/transactions", "Public (Portal Token)", "Header: X-Portal-Token", "{transactions: [{id, type, amount, date, receipt}]}"],

        # Audit Logs & Health (60-65)
        [60, "Audit System", "GET", "/api/v1/audit-logs", "ORG_ADMIN (AUDIT_VIEW)", "?branchId, ?action, ?page, ?limit", "{items: [{id, userId, userName, action, entity, timestamp}]}"],
        [61, "Audit System", "GET", "/api/v1/audit-logs/:id", "ORG_ADMIN (AUDIT_VIEW)", "Path: id", "{id, userId, action, entity, details: {}, timestamp}"],
        [62, "System & Health", "GET", "/api/v1/health", "Public", "None", "{status: 'OK', timestamp, database: 'CONNECTED'}"],
        [63, "System & Health", "GET", "/api/v1/version", "Public", "None", "{version: '1.0.0', contract: 'M0_V10_FROZEN'}"],
        [64, "System & Health", "GET", "/api/v1/metrics", "SUPER_ADMIN", "None", "{activeConnections, memoryUsage, uptime}"],
        [65, "System & Health", "POST", "/api/v1/system/seed-reset", "SUPER_ADMIN (Dev Only)", "{confirm: 'RESET_CONFIRM'}", "{message: 'Database re-seeded from TEST_DATA.json'}"]
    ]

    for row_idx, r in enumerate(api_routes, 5):
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        method = r[2]
        method_color = "166534" if method == "GET" else ("1E40AF" if method == "POST" else ("92400E" if method == "PUT" else "991B1B"))
        
        for col_idx, val in enumerate(r, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_data_font if col_idx in [1, 3] else (code_font if col_idx in [4, 6, 7] else data_font)
            cell.fill = fill
            cell.border = cell_border
            if col_idx in [1, 3]:
                cell.alignment = center_align
                if col_idx == 3:
                    cell.font = Font(name="Consolas", size=9, bold=True, color=method_color)
            else:
                cell.alignment = left_align
        ws6.row_dimensions[row_idx].height = 22

    auto_fit_columns(ws6, {"A": 10, "B": 22, "C": 14, "D": 38, "E": 28, "F": 45, "G": 55})

    # Save to disk
    output_filename = "d:/Flutter Money card/MONEY_CARD_DEVELOPMENT_ROADMAP_3-4_WEEKS.xlsx"
    wb.save(output_filename)
    print(f"Successfully generated: {output_filename}")

if __name__ == "__main__":
    create_roadmap_workbook()
